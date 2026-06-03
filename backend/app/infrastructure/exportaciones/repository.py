"""
Exportaciones — PDF y Excel.

Cubre:
  RF-ER-05  Reportes en formato PDF
  RF-ER-06  Reporte general de un paciente en PDF
  RF-ER-11  Exportar tablas a Excel
  RF-RB-06  Credencial de beneficiario en PDF
  RF-RB-07  Exportar lista filtrada de beneficiarios a Excel
  RF-SO-10  Comprobante de servicio / cita en PDF
  RF-PS-05  Contrato de comodato en PDF
"""
import io
import logging
from datetime import datetime, date
from pathlib import Path
from app.domain.exportaciones.ports import ExportacionesRepository
from app.domain.shared.current_user import CurrentUser
from app.domain.exportaciones.entities import FilePayload
from app.domain.exceptions import InternalError, NotFoundError, ValidationError
from app.infrastructure.persistence.oracle import get_db, rows_to_dicts, row_to_dict
from app.infrastructure.privacy.crypto import decrypt_row, PACIENTE_ENCRYPTED_FIELDS

logger = logging.getLogger(__name__)

_MSG_ERROR_INTERNO = 'Error interno del servidor'
_ORG_NAME = 'Asociación de Espina Bífida'
_COL_TOTAL_PACIENTES = 'Total Pacientes'
_COL_EDAD_PROMEDIO = 'Edad Promedio'
_COL_GENERO = 'Género'
_COL_ETAPA_VIDA = 'Etapa de Vida'
_COL_PAC_ATENDIDOS = 'Pacientes Atendidos'
_COL_METRICA = 'Métrica'

UPLOAD_DOCUMENTOS_DIR = Path(__file__).resolve().parents[3] / 'uploads' / 'documentos'
LOGO_PATH = Path(__file__).resolve().parents[3] / 'uploads' / 'logo.png'

def _strip(val):
    return val.strip() if isinstance(val, str) else val

def _date_str(val) -> str:
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y %H:%M')
    if isinstance(val, date):
        return val.strftime('%d/%m/%Y')
    return str(val)

def _paciente_str(paciente: dict, key: str) -> str:
    return paciente.get(key) or ''

def _paciente_strip_str(paciente: dict, key: str) -> str:
    return _strip(paciente.get(key)) or ''

def _usa_valvula_label(paciente: dict) -> str:
    return 'Sí' if _strip(paciente.get('usa_valvula')) == 'S' else 'No'

def _pdf_payload(buffer: io.BytesIO, filename: str) -> FilePayload:
    buffer.seek(0)
    return FilePayload(content=buffer.read(), media_type='application/pdf', filename=filename)

def _excel_payload(buffer: io.BytesIO, filename: str) -> FilePayload:
    buffer.seek(0)
    return FilePayload(content=buffer.read(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)

_MESES_N = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

_REPORTE_TIPO_LABELS = {
    'resumen': 'Reporte de Resumen de Período',
    'por-genero': 'Distribución por Género',
    'por-etapa-vida': 'Distribución por Etapa de Vida',
    'por-estado': 'Distribución por Estado de Residencia',
    'por-tipo-espina': 'Distribución por Tipo de Espina Bífida',
    'consolidado-mensual': 'Reporte Consolidado Mensual',
    'indicadores': 'Indicadores de Desempeño',
}

_INDICADORES_TABLA_CFGS = [
    ('por_curp', 'Sujetos por CURP', ['CURP N.L.', 'CURP Foráneo']),
    ('curp_nl_genero', 'CURP N.L. por Género', ['Hombre', 'Mujer']),
    ('curp_foraneo_genero', 'CURP Foráneo por Género', ['Hombre', 'Mujer']),
    ('residencia', 'Lugar de Residencia por Etapa', ['Viven en N.L.', 'Viven en otros estados']),
    ('nacimiento', 'País de Nacimiento por Etapa', ['Mexicanos', 'Nac. extranjera']),
    ('etapa_vida_genero', 'Etapa de Vida por Género', ['Hombre', 'Mujer']),
]

_REPORTE_SIMPLE_COL_NAMES = {
    'por-genero': _COL_GENERO,
    'por-etapa-vida': _COL_ETAPA_VIDA,
    'por-estado': 'Estado',
    'por-tipo-espina': 'Tipo de Espina',
}


def _reporte_pdf_pct(v, total):
    return f'{v / total * 100:.1f}%' if total else '—'


def _reporte_pdf_money(v):
    return f'${float(v or 0):,.2f}'


def _reporte_pdf_data_table(rows, widths, *, nav, navy_lt, border, row_alt, white, has_totals=False):
    from reportlab.platypus import Table, TableStyle
    t = Table(rows, colWidths=widths)
    cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), nav),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.4, border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2 if has_totals else -1), [white, row_alt]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
    ]
    if has_totals:
        cmds += [
            ('BACKGROUND', (0, -1), (-1, -1), navy_lt),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1, nav),
        ]
    t.setStyle(TableStyle(cmds))
    return t


def _reporte_pdf_kpi_table(headers, values, *, col, nav, navy_lt, border, white):
    from reportlab.platypus import Table, TableStyle
    n = len(headers)
    w = col / n
    t = Table([headers, values], colWidths=[w] * n)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), nav),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, 1), navy_lt),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, border),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    return t


def _reporte_pdf_page_footer(l_margin, r_margin, page_w, gray):
    from reportlab.lib.units import inch

    def _footer(cv, doc):
        cv.saveState()
        cv.setFont('Helvetica', 7)
        cv.setFillColor(gray)
        cv.drawString(l_margin, 0.4 * inch,
                      'Confidencial — Asociación de Espina Bífida de Nuevo León A.B.P.')
        cv.drawRightString(page_w - r_margin, 0.4 * inch, f'Pág. {doc.page}')
        cv.restoreState()
    return _footer


def _reporte_pdf_periodo_str(fecha_inicio, fecha_fin, mes, anio) -> str:
    if fecha_inicio or fecha_fin:
        return f'Período: {fecha_inicio or "—"} al {fecha_fin or "—"}   |   '
    if mes and anio:
        return f'Período: {_MESES_N[mes]} {anio}   |   '
    return ''


def _reporte_pdf_append_header(els, tipo, fecha_inicio, fecha_fin, mes, anio, *, h1, h2_style, note, nav):
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer, Image as RLImage, HRFlowable
    if LOGO_PATH.exists():
        logo = RLImage(str(LOGO_PATH), width=2.2 * inch, height=0.93 * inch)
        logo.hAlign = 'CENTER'
        els.append(logo)
        els.append(Spacer(1, 4))
    els.append(Paragraph(_ORG_NAME, h1))
    els.append(Paragraph(_REPORTE_TIPO_LABELS.get(tipo, f'Reporte: {tipo}'), h2_style))
    periodo_str = _reporte_pdf_periodo_str(fecha_inicio, fecha_fin, mes, anio)
    els.append(Paragraph(
        f'{periodo_str}Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', note))
    els.append(HRFlowable(width='100%', thickness=1, color=nav, spaceAfter=10, spaceBefore=6))


def _reporte_pdf_append_distrib_pct(els, h2, title, data, col_header, widths, *, theme):
    from reportlab.platypus import Paragraph, Spacer, KeepTogether
    if not data.get('labels'):
        return
    tot = data.get('total', 0)
    rows = [[col_header, 'Pacientes', '%']]
    for label, val in zip(data['labels'], data['values']):
        rows.append([label, str(val), _reporte_pdf_pct(val, tot)])
    rows.append(['Total', str(tot), '100%'])
    els.append(KeepTogether([
        Paragraph(title, h2),
        _reporte_pdf_data_table(rows, widths, has_totals=True, **theme),
    ]))
    els.append(Spacer(1, 8))


def _reporte_pdf_ciudades_rows(d_ciu: dict, top: int = 25) -> tuple[list, int]:
    """Build (rows_list, total) for the city-distribution PDF table."""
    ci_tot = d_ciu.get('total', 0)
    rows: list = [['Ciudad', 'Estado', 'Pacientes', '%']]
    for label, estado, val in zip(
        d_ciu['labels'][:top],
        d_ciu.get('estados', [''] * top)[:top],
        d_ciu['values'][:top],
    ):
        rows.append([label, estado, str(val), _reporte_pdf_pct(val, ci_tot)])
    if len(d_ciu['labels']) > top:
        resto = ci_tot - sum(d_ciu['values'][:top])
        rows.append(['Otras ciudades', '', str(resto), _reporte_pdf_pct(resto, ci_tot)])
    rows.append(['Total', '', str(ci_tot), '100%'])
    return rows, ci_tot


def _reporte_pdf_append_resumen(els, kwargs, fecha_inicio, fecha_fin, current_user, *, h2, col, theme):
    from app.infrastructure.reportes.repository import (
        reporte_resumen, reporte_por_genero, reporte_por_etapa_vida,
        reporte_por_estado, reporte_por_tipo_espina, reporte_servicios_por_tipo,
        reporte_estudios_por_tipo, reporte_pagos_exentos, reporte_por_ciudad,
    )
    from reportlab.platypus import Paragraph, Spacer, KeepTogether

    d_res = reporte_resumen(**kwargs)
    d_gen = reporte_por_genero(**kwargs)
    d_etapa = reporte_por_etapa_vida(**kwargs)
    d_esp = reporte_por_tipo_espina(**kwargs)
    d_est_r = reporte_por_estado(**kwargs)
    d_svc = reporte_servicios_por_tipo(
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)
    d_estu = reporte_estudios_por_tipo(
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)
    d_pag = reporte_pagos_exentos(
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)
    d_ciu = reporte_por_ciudad(current_user=current_user)

    total_p = d_res.get('total_pacientes', 0)
    els.append(Paragraph('Resumen Ejecutivo', h2))
    els.append(_reporte_pdf_kpi_table(
        [_COL_TOTAL_PACIENTES, 'Activos', 'Hombres', 'Mujeres', _COL_EDAD_PROMEDIO, 'Estados'],
        [str(total_p), str(d_res.get('activos', 0)),
         str(d_res.get('por_genero', {}).get('Hombre', 0)),
         str(d_res.get('por_genero', {}).get('Mujer', 0)),
         f'{d_res.get("edad_promedio", 0):.1f} años',
         str(d_res.get('estados_representados', 0))],
        col=col, **{k: theme[k] for k in ('nav', 'navy_lt', 'border', 'white')},
    ))
    els.append(Spacer(1, 10))

    _reporte_pdf_append_distrib_pct(
        els, h2, 'Distribución por Género', d_gen, _COL_GENERO,
        [col * .55, col * .22, col * .23], theme=theme)
    _reporte_pdf_append_distrib_pct(
        els, h2, 'Distribución por Etapa de Vida', d_etapa, _COL_ETAPA_VIDA,
        [col * .60, col * .20, col * .20], theme=theme)
    _reporte_pdf_append_distrib_pct(
        els, h2, 'Distribución por Tipo de Espina Bífida', d_esp, 'Tipo de Espina Bífida',
        [col * .60, col * .20, col * .20], theme=theme)

    if d_est_r.get('labels'):
        tot = d_est_r.get('total', 0)
        labels_e = d_est_r['labels'][:20]
        values_e = d_est_r['values'][:20]
        rows = [['Estado de Residencia', 'Pacientes', '%']]
        for label, val in zip(labels_e, values_e):
            rows.append([label, str(val), _reporte_pdf_pct(val, tot)])
        if len(d_est_r['labels']) > 20:
            resto = tot - sum(values_e)
            rows.append([
                f'Otros ({len(d_est_r["labels"]) - 20} estados)', str(resto),
                _reporte_pdf_pct(resto, tot),
            ])
        rows.append(['Total', str(tot), '100%'])
        els.append(KeepTogether([
            Paragraph('Distribución por Estado de Residencia', h2),
            _reporte_pdf_data_table(rows, [col * .58, col * .21, col * .21], has_totals=True, **theme),
        ]))
        els.append(Spacer(1, 8))

    if d_svc.get('labels'):
        montos = d_svc.get('montos', [0] * len(d_svc['labels']))
        rows = [['Servicio', 'Cantidad', 'Monto']]
        for label, val, monto in zip(d_svc['labels'], d_svc['values'], montos):
            rows.append([label, str(val), _reporte_pdf_money(monto)])
        rows.append(['Total', str(d_svc.get('total', 0)), _reporte_pdf_money(sum(montos))])
        els.append(KeepTogether([
            Paragraph('Servicios Brindados en el Período', h2),
            _reporte_pdf_data_table(rows, [col * .55, col * .18, col * .27], has_totals=True, **theme),
        ]))
        els.append(Spacer(1, 8))

    if d_estu.get('labels'):
        rows = [['Estudio / Servicio', 'Cantidad']]
        for label, val in zip(d_estu['labels'], d_estu['values']):
            rows.append([label, str(val)])
        rows.append(['Total', str(d_estu.get('total', 0))])
        els.append(KeepTogether([
            Paragraph('Estudios Realizados en el Período', h2),
            _reporte_pdf_data_table(rows, [col * .75, col * .25], has_totals=True, **theme),
        ]))
        els.append(Spacer(1, 8))

    els.append(Paragraph('Pagos Exentos vs Cuotas de Recuperación', h2))
    pg_rows = [
        ['Concepto', 'Cantidad', 'Monto Total'],
        ['Pagos Exentos', str(d_pag.get('total_exentos', 0)), _reporte_pdf_money(d_pag.get('monto_exentos', 0))],
        ['Cuotas de Recuperación', str(d_pag.get('total_cuotas', 0)), _reporte_pdf_money(d_pag.get('monto_cuotas', 0))],
        ['Total General',
         str(d_pag.get('total_exentos', 0) + d_pag.get('total_cuotas', 0)),
         _reporte_pdf_money(d_pag.get('monto_total', 0))],
    ]
    els.append(_reporte_pdf_data_table(pg_rows, [col * .55, col * .18, col * .27], has_totals=True, **theme))
    els.append(Spacer(1, 8))

    if d_ciu.get('labels'):
        top = 25
        rows, _ = _reporte_pdf_ciudades_rows(d_ciu, top)
        els.append(KeepTogether([
            Paragraph(f'Distribución por Ciudad de Residencia (Top {top})', h2),
            _reporte_pdf_data_table(
                rows, [col * .34, col * .29, col * .19, col * .18], has_totals=True, **theme),
        ]))


def _reporte_pdf_append_consolidado(els, mes, anio, current_user, *, h2, col, theme):
    from datetime import date as _date
    from app.infrastructure.reportes.repository import reporte_consolidado_mensual
    from reportlab.platypus import Paragraph, Spacer, KeepTogether

    _mes = mes or _date.today().month
    _anio = anio or _date.today().year
    d = reporte_consolidado_mensual(mes=_mes, anio=_anio, current_user=current_user)

    els.append(Paragraph(f'Período: {_MESES_N[_mes]} {_anio}', h2))
    els.append(_reporte_pdf_kpi_table(
        [_COL_PAC_ATENDIDOS, 'Total Servicios', 'Monto Servicios', 'Total Ventas', 'Monto Ventas'],
        [str(d.get('pacientes_atendidos', 0)), str(d.get('total_servicios', 0)),
         _reporte_pdf_money(d.get('monto_servicios', 0)), str(d.get('total_ventas', 0)),
         _reporte_pdf_money(d.get('monto_ventas', 0))],
        col=col, **{k: theme[k] for k in ('nav', 'navy_lt', 'border', 'white')},
    ))
    els.append(Spacer(1, 12))

    citas_est = d.get('citas_por_estatus', {})
    if citas_est:
        rows = [['Estatus', 'Cantidad']]
        tot_c = sum(citas_est.values())
        for k, v in citas_est.items():
            rows.append([k, str(v)])
        rows.append(['Total', str(tot_c)])
        els.append(KeepTogether([
            Paragraph('Citas por Estatus', h2),
            _reporte_pdf_data_table(rows, [col * .65, col * .35], has_totals=True, **theme),
        ]))
        els.append(Spacer(1, 8))

    pg = d.get('por_genero', {})
    if pg:
        rows = [[_COL_GENERO, _COL_PAC_ATENDIDOS]]
        tot_g = sum(pg.values())
        for k, v in pg.items():
            rows.append([k, str(v)])
        rows.append(['Total', str(tot_g)])
        els.append(KeepTogether([
            Paragraph('Pacientes Atendidos por Género', h2),
            _reporte_pdf_data_table(rows, [col * .65, col * .35], has_totals=True, **theme),
        ]))


def _reporte_pdf_append_indicadores(els, fecha_inicio, fecha_fin, current_user, *, h2, col, theme):
    from app.infrastructure.reportes.repository import indicadores_desempeno
    from reportlab.platypus import Paragraph, Spacer, KeepTogether

    d = indicadores_desempeno(
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)

    els.append(_reporte_pdf_kpi_table(
        ['Beneficiarios Activos', 'Nuevos en Período', 'Hombres', 'Mujeres'],
        [str(d.get('beneficiarios_activos', 0)), str(d.get('nuevos_en_periodo', 0)),
         str(d.get('hombres', 0)), str(d.get('mujeres', 0))],
        col=col, **{k: theme[k] for k in ('nav', 'navy_lt', 'border', 'white')},
    ))
    els.append(Spacer(1, 10))

    municipios = d.get('municipios', [])
    if municipios:
        mun_tot = sum(m.get('value', 0) for m in municipios)
        rows = [['Municipio / Lugar', 'Beneficiarios', '%']]
        for m in municipios:
            rows.append([
                m.get('label', ''), str(m.get('value', 0)),
                _reporte_pdf_pct(m.get('value', 0), mun_tot),
            ])
        rows.append(['Total', str(mun_tot), '100%'])
        els.append(KeepTogether([
            Paragraph('Beneficiarios por Municipio (Nuevo León)', h2),
            _reporte_pdf_data_table(rows, [col * .58, col * .22, col * .20], has_totals=True, **theme),
        ]))
        els.append(Spacer(1, 10))

    tablas = d.get('tablas', {})
    for key, titulo_t, cols in _INDICADORES_TABLA_CFGS:
        t_rows = tablas.get(key, [])
        if not t_rows:
            continue
        rows = [[_COL_ETAPA_VIDA, cols[0], cols[1], 'Total']]
        for r in t_rows:
            rows.append([
                r.get('etapa', ''),
                str(r.get(cols[0], 0) or 0),
                str(r.get(cols[1], 0) or 0),
                str(r.get('total', 0) or 0),
            ])
        els.append(KeepTogether([
            Paragraph(titulo_t, h2),
            _reporte_pdf_data_table(
                rows, [col * .44, col * .18, col * .20, col * .18], has_totals=True, **theme),
        ]))
        els.append(Spacer(1, 8))


def _reporte_pdf_append_simple(els, tipo, kwargs, fecha_inicio, fecha_fin, *, col, theme):
    from app.infrastructure.reportes.repository import (
        reporte_por_genero, reporte_por_etapa_vida, reporte_por_estado,
        reporte_por_tipo_espina, reporte_resumen,
    )
    from reportlab.platypus import Table, TableStyle, Spacer

    report_funcs = {
        'por-genero': reporte_por_genero,
        'por-etapa-vida': reporte_por_etapa_vida,
        'por-estado': reporte_por_estado,
        'por-tipo-espina': reporte_por_tipo_espina,
    }
    func = report_funcs.get(tipo)
    if not func:
        raise ValidationError(f'Tipo de reporte no válido: {tipo}')
    data = func(**kwargs)
    d_res = reporte_resumen(**kwargs)

    ctx = Table(
        [['Pacientes Activos', 'Período', 'Generado'],
         [str(d_res.get('activos', 0)),
          f'{fecha_inicio or "—"} al {fecha_fin or "—"}',
          datetime.now().strftime('%d/%m/%Y')]],
        colWidths=[col / 3] * 3,
    )
    ctx.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), theme['nav']),
        ('TEXTCOLOR', (0, 0), (-1, 0), theme['white']),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, 1), theme['navy_lt']),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, theme['border']),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    els.append(ctx)
    els.append(Spacer(1, 12))

    tot = data.get('total', 0)
    rows = [[_REPORTE_SIMPLE_COL_NAMES.get(tipo, 'Categoría'), 'Cantidad', '%']]
    for label, val in zip(data['labels'], data['values']):
        rows.append([label, str(val), _reporte_pdf_pct(val, tot)])
    rows.append(['Total', str(tot), '100%'])
    els.append(_reporte_pdf_data_table(
        rows, [col * .60, col * .20, col * .20], has_totals=True, **theme))


def _reporte_pdf_filename(tipo, mes, anio, fecha_inicio, fecha_fin) -> str:
    ts = datetime.now().strftime('%Y%m%d')
    if tipo == 'consolidado-mensual':
        return f'consolidado_{mes or ""}_{anio or ""}_{ts}.pdf'
    if tipo == 'indicadores':
        return f'indicadores_{fecha_inicio or ts}_{fecha_fin or ts}.pdf'
    return f'reporte_{tipo}_{fecha_inicio or ts}.pdf'


def _exportar_reporte_pdf(
    tipo: str = 'resumen',
    genero: str | None = None,
    estado: str | None = None,
    tipo_espina: int | None = None,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    mes: int | None = None,
    anio: int | None = None,
    current_user: CurrentUser | None = None,
):
    """Generar reporte estadístico en PDF (RF-ER-05)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate

    nav = colors.HexColor('#1e3a5f')
    navy_lt = colors.HexColor('#f0f4f8')
    border = colors.HexColor('#e2e8f0')
    gray = colors.HexColor('#64748b')
    row_alt = colors.HexColor('#f8fafc')
    white = colors.white
    theme = {'nav': nav, 'navy_lt': navy_lt, 'border': border, 'row_alt': row_alt, 'white': white}

    page_w, _ = letter
    l_margin = r_margin = 0.75 * inch
    col = page_w - l_margin - r_margin

    try:
        styles = getSampleStyleSheet()
        h1 = ParagraphStyle('H1', parent=styles['Title'], fontSize=16, leading=20, spaceAfter=2)
        h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=11, textColor=nav,
                            spaceBefore=14, spaceAfter=4)
        note = ParagraphStyle('Note', parent=styles['Normal'], fontSize=8, textColor=gray)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=letter,
            topMargin=0.6 * inch, bottomMargin=0.75 * inch,
            leftMargin=l_margin, rightMargin=r_margin,
        )
        els = []
        _reporte_pdf_append_header(
            els, tipo, fecha_inicio, fecha_fin, mes, anio,
            h1=h1, h2_style=styles['Heading2'], note=note, nav=nav)

        kwargs = {
            'genero': genero, 'estado': estado, 'tipo_espina': tipo_espina,
            'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'current_user': current_user,
        }
        if tipo == 'resumen':
            _reporte_pdf_append_resumen(
                els, kwargs, fecha_inicio, fecha_fin, current_user,
                h2=h2, col=col, theme=theme)
        elif tipo == 'consolidado-mensual':
            _reporte_pdf_append_consolidado(
                els, mes, anio, current_user, h2=h2, col=col, theme=theme)
        elif tipo == 'indicadores':
            _reporte_pdf_append_indicadores(
                els, fecha_inicio, fecha_fin, current_user, h2=h2, col=col, theme=theme)
        else:
            _reporte_pdf_append_simple(
                els, tipo, kwargs, fecha_inicio, fecha_fin, col=col, theme=theme)

        page_footer = _reporte_pdf_page_footer(l_margin, r_margin, page_w, gray)
        doc.build(els, onFirstPage=page_footer, onLaterPages=page_footer)
        return _pdf_payload(buf, _reporte_pdf_filename(tipo, mes, anio, fecha_inicio, fecha_fin))

    except (NotFoundError, ValidationError, InternalError):
        raise
    except Exception:
        logger.exception('Error al generar PDF de reporte')
        raise InternalError(_MSG_ERROR_INTERNO)

def _beneficiario_pdf_fields(paciente: dict, nombre: str) -> list[tuple]:
    """Return ordered (label, value) pairs for a beneficiary expediente PDF."""
    return [
        ('Nombre', nombre),
        ('CURP', _paciente_str(paciente, 'curp')),
        (_COL_GENERO, _paciente_strip_str(paciente, 'genero')),
        ('Fecha de Nacimiento', _date_str(paciente.get('fecha_nacimiento'))),
        ('Tipo de Sangre', _paciente_str(paciente, 'tipo_sangre')),
        ('Usa Válvula', _usa_valvula_label(paciente)),
        ('Dirección', _paciente_str(paciente, 'direccion')),
        ('Colonia', _paciente_str(paciente, 'colonia')),
        ('Ciudad', _paciente_str(paciente, 'ciudad')),
        ('Estado', _paciente_str(paciente, 'estado')),
        ('C.P.', _paciente_str(paciente, 'codigo_postal')),
        ('Teléfono Casa', _paciente_str(paciente, 'telefono_casa')),
        ('Teléfono Celular', _paciente_str(paciente, 'telefono_celular')),
        ('Correo', _paciente_str(paciente, 'correo_electronico')),
        ('Contacto Emergencia', _paciente_str(paciente, 'en_emergencia_avisar_a')),
        ('Tel. Emergencia', _paciente_str(paciente, 'telefono_emergencia')),
        ('Membresía', _paciente_strip_str(paciente, 'membresia_estatus')),
        ('Tipo Cuota', _paciente_strip_str(paciente, 'tipo_cuota')),
        ('Fecha Alta', _date_str(paciente.get('fecha_alta'))),
    ]


def _exportar_beneficiario_pdf(folio: str, _current_user: CurrentUser | None = None):
    """Generar reporte PDF de un beneficiario con sus datos y documentos (RF-ER-06)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute('SELECT * FROM PACIENTE WHERE FOLIO = :folio', {'folio': folio})
            paciente = row_to_dict(cur)
            if not paciente:
                raise NotFoundError('Beneficiario no encontrado')
            paciente = decrypt_row(paciente, PACIENTE_ENCRYPTED_FIELDS)
            nombre = f'{paciente["nombre"]} {paciente["apellido_paterno"]} {paciente.get("apellido_materno") or ""}'.strip()
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []
        if LOGO_PATH.exists():
            logo = RLImage(str(LOGO_PATH), width=2.5*inch, height=1.05*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 6))
        elements.append(Paragraph(_ORG_NAME, styles['Title']))
        elements.append(Paragraph(f'Expediente del Beneficiario — {nombre}', styles['Heading2']))
        elements.append(Paragraph(f'Folio: {folio} | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
        elements.append(Spacer(1, 20))
        table_data = [['Campo', 'Valor']]
        for label, val in _beneficiario_pdf_fields(paciente, nombre):
            table_data.append([label, str(val)])
        t = Table(table_data, colWidths=[180, 300])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)
        return _pdf_payload(buf, f'expediente_{folio}_{datetime.now().strftime("%Y%m%d")}.pdf')
    except (NotFoundError, ValidationError, InternalError):
        raise
    except Exception:
        logger.exception('Error al generar PDF del beneficiario')
        raise InternalError(_MSG_ERROR_INTERNO)

_CREDENCIAL_IMAGE_FORMATS = frozenset({'JPG', 'JPEG', 'PNG', 'WEBP'})
_CREDENCIAL_FOTO_MARKERS = ('foto', 'fotografia', 'imagen')


def _credencial_pdf_sv(paciente: dict, key: str) -> str:
    return _strip(paciente.get(key)) or ''


def _credencial_pdf_lbl(cv, x, y, txt, navy):
    cv.setFillColor(navy)
    cv.setFont('Helvetica-Bold', 6.5)
    cv.drawString(x, y, txt.upper())


def _credencial_pdf_val(cv, x, y, txt, black, bold=False, size=8):
    cv.setFillColor(black)
    cv.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
    cv.drawString(x, y, str(txt) if txt else '-')


def _credencial_pdf_load_data(folio: str):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute('SELECT * FROM PACIENTE WHERE FOLIO = :folio', {'folio': folio})
        paciente = row_to_dict(cur)
        if not paciente:
            raise NotFoundError('Beneficiario no encontrado')
        paciente = decrypt_row(paciente, PACIENTE_ENCRYPTED_FIELDS)
        cur.execute(
            'SELECT te.NOMBRE FROM PACIENTE_TIPO_ESPINA pte '
            'JOIN TIPO_ESPINA_BIFIDA te ON te.ID_TIPO_ESPINA = pte.ID_TIPO_ESPINA '
            'WHERE pte.ID_PACIENTE = :id', {'id': paciente['id_paciente']}
        )
        tipos = [r[0].strip() for r in cur.fetchall()]
        cur.execute(
            'SELECT dp.ID_DOCUMENTO, dp.RUTA_ARCHIVO, dp.FORMATO_ARCHIVO, dp.FECHA_CARGA, '
            '       td.NOMBRE AS TIPO_NOMBRE '
            'FROM DOCUMENTO_PACIENTE dp '
            'LEFT JOIN TIPO_DOCUMENTO td ON td.ID_TIPO_DOCUMENTO = dp.ID_TIPO_DOCUMENTO '
            "WHERE dp.ID_PACIENTE = :id AND dp.ACTIVO = 'S'",
            {'id': paciente['id_paciente']}
        )
        documentos = rows_to_dicts(cur)
    return paciente, tipos, documentos


def _credencial_pdf_resolve_foto_path(documentos: list) -> Path | None:
    imagenes = [
        d for d in documentos
        if str(d.get('formato_archivo') or '').strip().upper() in _CREDENCIAL_IMAGE_FORMATS
    ]
    if not imagenes:
        return None
    fotos_por_tipo = [
        d for d in imagenes
        if any(
            marker in str(d.get('tipo_nombre') or '').lower()
            for marker in _CREDENCIAL_FOTO_MARKERS
        )
    ]
    candidatas = fotos_por_tipo or imagenes
    candidatas.sort(key=lambda d: d.get('fecha_carga') or datetime.min, reverse=True)
    ruta = str(candidatas[0].get('ruta_archivo') or '').strip()
    if not ruta:
        return None
    ruta_local = (UPLOAD_DOCUMENTOS_DIR / Path(ruta).name).resolve()
    if ruta_local.exists() and ruta_local.is_file():
        return ruta_local
    return None


def _credencial_pdf_display_fields(paciente: dict, tipos: list) -> dict:
    sv = lambda k: _credencial_pdf_sv(paciente, k)
    estado = sv('estado')
    return {
        'sv': sv,
        'nombre_completo': f"{sv('nombre')} {sv('apellido_paterno')} {sv('apellido_materno')}".strip(),
        'direccion': f"{sv('direccion')}, {sv('colonia')}",
        'ciudad_est': f"{sv('ciudad')}{', ' + estado if estado else ''}",
        'fecha_nac': str(sv('fecha_nacimiento') or '')[:10],
        'fecha_exp': str(sv('fecha_alta') or '')[:10],
        'valvula': 'Si' if (sv('usa_valvula') or 'N').upper() == 'S' else 'No',
        'padecimiento': ', '.join(tipos) if tipos else 'No especificado',
    }


def _credencial_pdf_draw_photo_placeholder(cv, photo_x, photo_y, photo_w, photo_h, gray, cm):
    cv.setFillColor(gray)
    cv.setFont('Helvetica', 6)
    cv.drawCentredString(photo_x + photo_w / 2, photo_y + photo_h / 2 + 0.3 * cm, 'FOTOGRAFIA')
    cv.drawCentredString(photo_x + photo_w / 2, photo_y + photo_h / 2 - 0.1 * cm, 'DEL PORTADOR')


def _credencial_pdf_draw_header(cv, w, h, hdr_h, folio, fields, theme, cm):
    navy, white = theme['navy'], theme['white']
    cv.setFillColor(navy)
    cv.rect(0, h - hdr_h, w, hdr_h, fill=1, stroke=0)
    logo_w_hdr = 1.4 * cm
    logo_h_hdr = 1.4 * cm
    if LOGO_PATH.exists():
        try:
            cv.drawImage(
                str(LOGO_PATH), 0.2 * cm, h - hdr_h + 0.1 * cm,
                logo_w_hdr, logo_h_hdr, preserveAspectRatio=True, anchor='c', mask='auto')
        except Exception:
            pass
    txt_x = logo_w_hdr + 0.45 * cm
    cv.setFillColor(white)
    cv.setFont('Helvetica-Bold', 9)
    cv.drawString(txt_x, h - 1.05 * cm, 'ESPINA BIFIDA - Asociacion de Nuevo Leon ABP')
    cv.setFont('Helvetica', 7)
    cv.drawString(txt_x, h - 1.4 * cm, 'CREDENCIAL DE BENEFICIARIO')
    cv.setFont('Helvetica-Bold', 9)
    cv.drawRightString(w - 0.5 * cm, h - 1.05 * cm, f'Folio: {folio}')
    cv.setFont('Helvetica', 7)
    membresia = fields['sv']('id_paciente') or 'N/A'
    cv.drawRightString(w - 0.5 * cm, h - 1.4 * cm, f'Membresia No.: {membresia}')


def _credencial_pdf_draw_photo(cv, _w, h, hdr_h, foto_path, theme, cm):
    from reportlab.lib import colors
    lgray = theme['lgray']
    gray = theme['gray']
    photo_x, photo_y = 0.4 * cm, h - hdr_h - 3.9 * cm
    photo_w, photo_h = 2.8 * cm, 3.5 * cm
    cv.setFillColor(lgray)
    cv.setStrokeColor(colors.HexColor('#cbd5e1'))
    cv.rect(photo_x, photo_y, photo_w, photo_h, fill=1, stroke=1)
    if foto_path:
        try:
            cv.drawImage(
                str(foto_path),
                photo_x + 0.05 * cm, photo_y + 0.05 * cm,
                photo_w - 0.10 * cm, photo_h - 0.10 * cm,
                preserveAspectRatio=True, anchor='c', mask='auto',
            )
            return photo_y
        except Exception:
            logger.warning(
                'No se pudo renderizar la foto del beneficiario en credencial', exc_info=True)
    _credencial_pdf_draw_photo_placeholder(cv, photo_x, photo_y, photo_w, photo_h, gray, cm)
    return photo_y


def _credencial_pdf_draw_left_column(cv, photo_y, fields, theme, cm):
    navy, black = theme['navy'], theme['black']
    sv = fields['sv']
    lx = 0.4 * cm
    ly = photo_y - 0.65 * cm
    _credencial_pdf_lbl(cv, lx, ly, 'Nombre Completo', navy)
    ly -= 0.35 * cm
    _credencial_pdf_val(cv, lx, ly, fields['nombre_completo'][:42], black, bold=True, size=7)
    ly -= 0.5 * cm
    _credencial_pdf_lbl(cv, lx, ly, 'Direccion', navy)
    ly -= 0.32 * cm
    _credencial_pdf_val(cv, lx, ly, fields['direccion'][:38], black, size=7)
    ly -= 0.28 * cm
    _credencial_pdf_val(cv, lx, ly, fields['ciudad_est'][:38], black, size=7)
    ly -= 0.45 * cm
    _credencial_pdf_lbl(cv, lx, ly, 'Tel. Casa', navy)
    ly -= 0.32 * cm
    _credencial_pdf_val(cv, lx, ly, sv('telefono_casa') or '-', black, size=7)
    ly -= 0.45 * cm
    _credencial_pdf_lbl(cv, lx, ly, 'Nombre de Padre / Madre', navy)
    ly -= 0.32 * cm
    _credencial_pdf_val(cv, lx, ly, (sv('nombre_padre_madre') or '-')[:35], black, size=7)
    ly -= 0.45 * cm
    _credencial_pdf_lbl(cv, lx, ly, 'Fecha de Expedicion', navy)
    ly -= 0.32 * cm
    _credencial_pdf_val(cv, lx, ly, fields['fecha_exp'] or '-', black, size=7)


def _credencial_pdf_draw_right_column(cv, w, h, hdr_h, fields, theme, cm):
    from reportlab.lib import colors
    navy, navy_l, black = theme['navy'], theme['navy_l'], theme['black']
    sv = fields['sv']
    mid_x = 3.6 * cm
    cv.setStrokeColor(colors.HexColor('#e2e8f0'))
    cv.line(mid_x, 0.55 * cm, mid_x, h - hdr_h - 0.2 * cm)
    rx = mid_x + 0.4 * cm
    ry = h - hdr_h - 0.55 * cm
    _credencial_pdf_lbl(cv, rx, ry, 'Padecimiento', navy)
    ry -= 0.32 * cm
    _credencial_pdf_val(cv, rx, ry, fields['padecimiento'][:55], black, bold=True, size=7)
    ry -= 0.5 * cm
    col2 = rx + 3.5 * cm
    _credencial_pdf_lbl(cv, rx, ry, 'Tipo de Sangre', navy)
    _credencial_pdf_lbl(cv, col2, ry, 'Tiene Valvula', navy)
    ry -= 0.32 * cm
    _credencial_pdf_val(cv, rx, ry, sv('tipo_sangre') or '-', black, bold=True, size=9)
    _credencial_pdf_val(cv, col2, ry, fields['valvula'], black, bold=True, size=8)
    ry -= 0.55 * cm
    _credencial_pdf_lbl(cv, rx, ry, 'En caso de accidente avisar a', navy)
    ry -= 0.32 * cm
    _credencial_pdf_val(cv, rx, ry, (sv('en_emergencia_avisar_a') or '-')[:42], black, size=7)
    ry -= 0.42 * cm
    _credencial_pdf_lbl(cv, rx, ry, 'Telefono de Emergencia', navy)
    ry -= 0.32 * cm
    _credencial_pdf_val(cv, rx, ry, sv('telefono_emergencia') or '-', black, size=7)
    ry -= 0.42 * cm
    _credencial_pdf_lbl(cv, rx, ry, 'Correo Electronico', navy)
    ry -= 0.32 * cm
    _credencial_pdf_val(cv, rx, ry, (sv('correo_electronico') or '-')[:45], black, size=7)
    ry -= 0.52 * cm
    cv.setStrokeColor(colors.HexColor('#e2e8f0'))
    cv.line(rx, ry + 0.15 * cm, w - 0.4 * cm, ry + 0.15 * cm)
    ry -= 0.25 * cm
    c3a = rx
    c3b = rx + 2.2 * cm
    c3c = rx + 4.6 * cm
    _credencial_pdf_lbl(cv, c3a, ry, 'Datos de Nacimiento', navy)
    ry -= 0.32 * cm
    _credencial_pdf_lbl(cv, c3a, ry, 'Fecha', navy)
    _credencial_pdf_lbl(cv, c3b, ry, 'Lugar Nac.', navy)
    _credencial_pdf_lbl(cv, c3c, ry, 'Hospital', navy)
    ry -= 0.32 * cm
    _credencial_pdf_val(cv, c3a, ry, fields['fecha_nac'] or '-', black, size=7)
    _credencial_pdf_val(cv, c3b, ry, (sv('estado_nacimiento') or '-')[:16], black, size=7)
    _credencial_pdf_val(cv, c3c, ry, (sv('hospital_nacimiento') or '-')[:18], black, size=7)
    ry -= 0.5 * cm
    box_w = w - rx - 0.4 * cm
    box_h = 0.9 * cm
    cv.setFillColor(navy_l)
    cv.setStrokeColor(navy)
    cv.roundRect(rx, ry - box_h + 0.25 * cm, box_w, box_h, 0.15 * cm, fill=1, stroke=1)
    cx_box = rx + box_w / 2
    cv.setFillColor(navy)
    cv.setFont('Helvetica-Bold', 6.5)
    cv.drawCentredString(cx_box, ry - 0.08 * cm, 'ASOCIACION DE ESPINA BIFIDA DE NUEVO LEON ABP')
    cv.setFont('Helvetica', 6)
    cv.drawCentredString(cx_box, ry - 0.42 * cm, 'www.espinabifida.org.mx')


def _credencial_pdf_draw_footer(cv, w, fields, theme, cm):
    lgray, gray = theme['lgray'], theme['gray']
    sv = fields['sv']
    cv.setFillColor(lgray)
    cv.rect(0, 0, w, 0.55 * cm, fill=1, stroke=0)
    cv.setFillColor(gray)
    cv.setFont('Helvetica', 6)
    cuota = sv('tipo_cuota') or 'No asignada'
    vencimiento = str(sv('fecha_vencimiento_membresia') or '')[:10] or 'Indefinida'
    cv.drawString(0.4 * cm, 0.2 * cm, f'Cuota: {cuota}')
    cv.drawCentredString(w / 2, 0.2 * cm, f'Vigencia: {vencimiento}')
    cv.drawRightString(w - 0.4 * cm, 0.2 * cm, f'CURP: {sv("curp") or "N/A"}')


def _credencial_pdf_render(cv, w, h, folio, paciente, tipos, foto_path):
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    theme = {
        'navy': colors.HexColor('#00328b'),
        'navy_l': colors.HexColor('#e8eef8'),
        'gray': colors.HexColor('#64748b'),
        'lgray': colors.HexColor('#f1f5f9'),
        'black': colors.HexColor('#1e293b'),
        'white': colors.white,
    }
    fields = _credencial_pdf_display_fields(paciente, tipos)
    hdr_h = 1.6 * cm
    cv.setFillColor(theme['white'])
    cv.rect(0, 0, w, h, fill=1, stroke=0)
    _credencial_pdf_draw_header(cv, w, h, hdr_h, folio, fields, theme, cm)
    photo_y = _credencial_pdf_draw_photo(cv, w, h, hdr_h, foto_path, theme, cm)
    _credencial_pdf_draw_left_column(cv, photo_y, fields, theme, cm)
    _credencial_pdf_draw_right_column(cv, w, h, hdr_h, fields, theme, cm)
    _credencial_pdf_draw_footer(cv, w, fields, theme, cm)


def _exportar_credencial_pdf(folio: str, _current_user: CurrentUser | None = None):
    """Generar credencial del beneficiario en PDF (RF-RB-06)."""
    from reportlab.lib.pagesizes import landscape, A6
    from reportlab.pdfgen import canvas as pdf_canvas
    try:
        paciente, tipos, documentos = _credencial_pdf_load_data(folio)
        foto_path = _credencial_pdf_resolve_foto_path(documentos)
        w, h = landscape(A6)
        buf = io.BytesIO()
        cv = pdf_canvas.Canvas(buf, pagesize=(w, h))
        _credencial_pdf_render(cv, w, h, folio, paciente, tipos, foto_path)
        cv.save()
        return _pdf_payload(buf, f'credencial_{folio}.pdf')
    except (NotFoundError, ValidationError, InternalError):
        raise
    except Exception:
        logger.exception('Error al generar credencial PDF')
        raise InternalError(_MSG_ERROR_INTERNO)


def _exportar_comprobante_cita(id_cita: int, _current_user: CurrentUser | None = None):
    """Generar comprobante PDF de una cita con sus servicios (RF-SO-10)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT c.ID_CITA, c.FECHA_HORA, c.ESTATUS, c.NOTAS,\n                          p.NOMBRE || ' ' || p.APELLIDO_PATERNO || ' ' || NVL(p.APELLIDO_MATERNO, '') AS nombre_paciente,\n                          p.FOLIO AS folio_paciente\n                   FROM CITA c\n                   JOIN PACIENTE p ON p.ID_PACIENTE = c.ID_PACIENTE\n                   WHERE c.ID_CITA = :id_cita", {'id_cita': id_cita})
            cita = row_to_dict(cur)
            if not cita:
                raise NotFoundError('Cita no encontrada')
            cur.execute('SELECT s.NOMBRE, d.CANTIDAD, d.MONTO_PAGADO, d.CANCELADO\n                   FROM DETALLE_CITA_SERVICIO d\n                   JOIN SERVICIO s ON s.ID_SERVICIO = d.ID_SERVICIO\n                   WHERE d.ID_CITA = :id_cita', {'id_cita': id_cita})
            servicios = rows_to_dicts(cur)
            cur.execute(
                'SELECT DISTINCT dr.NOMBRE || \' \' || dr.APELLIDO_PATERNO AS nombre_doctor,'
                ' dr.ESPECIALIDAD'
                ' FROM DETALLE_CITA_SERVICIO d'
                ' JOIN DOCTOR dr ON dr.ID_DOCTOR = d.ID_DOCTOR'
                ' WHERE d.ID_CITA = :id_cita AND d.ID_DOCTOR IS NOT NULL',
                {'id_cita': id_cita}
            )
            doctores = rows_to_dicts(cur)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []
        if LOGO_PATH.exists():
            logo = RLImage(str(LOGO_PATH), width=2.5*inch, height=1.05*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 6))
        elements.append(Paragraph(_ORG_NAME, styles['Title']))
        elements.append(Paragraph('Comprobante de Servicio', styles['Heading2']))
        elements.append(Spacer(1, 10))
        info_data = [['Paciente', cita['nombre_paciente']], ['Folio Paciente', cita['folio_paciente']], ['Fecha', _date_str(cita['fecha_hora'])], ['Estatus', _strip(cita['estatus'])], ['Notas', cita.get('notas') or '']]
        if doctores:
            for doc_row in doctores:
                info_data.append(['Doctor', f'{doc_row['nombre_doctor']} — {_strip(doc_row.get('especialidad') or '')}'])
        t_info = Table(info_data, colWidths=[120, 360])
        t_info.setStyle(TableStyle([('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 10), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4)]))
        elements.append(t_info)
        elements.append(Spacer(1, 15))
        elements.append(Paragraph('Servicios', styles['Heading3']))
        svc_data = [['Servicio', 'Cantidad', 'Monto', 'Cancelado']]
        total_monto = 0.0
        for s in servicios:
            monto = float(s.get('monto_pagado') or 0)
            total_monto += monto
            svc_data.append([_strip(s['nombre']), str(s.get('cantidad', 1)), f'${monto:,.2f}', 'Sí' if _strip(s.get('cancelado')) == 'S' else 'No'])
        svc_data.append(['', '', f'Total: ${total_monto:,.2f}', ''])
        t_svc = Table(svc_data, colWidths=[200, 80, 100, 100])
        t_svc.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4)]))
        elements.append(t_svc)
        doc.build(elements)
        return _pdf_payload(buf, f'comprobante_cita_{id_cita}.pdf')
    except (NotFoundError, ValidationError, InternalError):
        raise
    except Exception:
        logger.exception('Error al generar comprobante PDF')
        raise InternalError(_MSG_ERROR_INTERNO)

def _exportar_contrato_comodato(id_comodato: int, _current_user: CurrentUser | None = None):
    """Generar contrato de comodato en PDF (RF-PS-05)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT c.*, p.NOMBRE || ' ' || p.APELLIDO_PATERNO || ' ' || NVL(p.APELLIDO_MATERNO, '') AS nombre_paciente,\n                          p.FOLIO AS folio_paciente, pr.NOMBRE AS nombre_equipo,\n                          eq.NUMERO_SERIE, eq.MARCA, eq.MODELO\n                   FROM COMODATO c\n                   JOIN PACIENTE p ON p.ID_PACIENTE = c.ID_PACIENTE\n                   JOIN PRODUCTO pr ON pr.ID_PRODUCTO = c.ID_EQUIPO\n                   LEFT JOIN EQUIPO_MEDICO eq ON eq.ID_PRODUCTO = c.ID_EQUIPO\n                   WHERE c.ID_COMODATO = :id", {'id': id_comodato})
            com = row_to_dict(cur)
            if not com:
                raise NotFoundError('Comodato no encontrado')
            cur.execute('SELECT * FROM PACIENTE WHERE ID_PACIENTE = :id', {'id': com['id_paciente']})
            paciente = row_to_dict(cur)
            paciente = decrypt_row(paciente, PACIENTE_ENCRYPTED_FIELDS)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        styles = getSampleStyleSheet()
        elements = []
        if LOGO_PATH.exists():
            logo = RLImage(str(LOGO_PATH), width=2.5*inch, height=1.05*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 6))
        elements.append(Paragraph(_ORG_NAME, styles['Title']))
        elements.append(Paragraph('CONTRATO DE COMODATO', styles['Heading2']))
        elements.append(Spacer(1, 15))
        folio_com = _strip(com.get('folio_comodato')) or str(id_comodato)
        elements.append(Paragraph(f'<b>Folio:</b> {folio_com}', styles['Normal']))
        elements.append(Paragraph(f'<b>Fecha:</b> {_date_str(com.get('fecha_prestamo'))}', styles['Normal']))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph('<b>DATOS DEL BENEFICIARIO</b>', styles['Heading3']))
        ben_data = [['Nombre', com['nombre_paciente']], ['Folio', com['folio_paciente']], ['Dirección', paciente.get('direccion') or ''], ['Teléfono', paciente.get('telefono_celular') or paciente.get('telefono_casa') or '']]
        t_ben = Table(ben_data, colWidths=[120, 360])
        t_ben.setStyle(TableStyle([('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 10), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]))
        elements.append(t_ben)
        elements.append(Spacer(1, 10))
        elements.append(Paragraph('<b>EQUIPO EN PRÉSTAMO</b>', styles['Heading3']))
        eq_data = [['Equipo', _strip(com.get('nombre_equipo')) or ''], ['Número de Serie', _strip(com.get('numero_serie')) or 'N/A'], ['Marca', _strip(com.get('marca')) or 'N/A'], ['Modelo', _strip(com.get('modelo')) or 'N/A']]
        t_eq = Table(eq_data, colWidths=[120, 360])
        t_eq.setStyle(TableStyle([('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 10), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]))
        elements.append(t_eq)
        elements.append(Spacer(1, 10))
        elements.append(Paragraph('<b>CONDICIONES ECONÓMICAS</b>', styles['Heading3']))
        monto_total = float(com.get('monto_total') or 0)
        monto_pagado = float(com.get('monto_pagado') or 0)
        saldo = float(com.get('saldo_pendiente') or 0)
        exento = _strip(com.get('exento_pago'))
        fin_data = [['Monto Total', f'${monto_total:,.2f}'], ['Monto Pagado', f'${monto_pagado:,.2f}'], ['Saldo Pendiente', f'${saldo:,.2f}'], ['Exento de Pago', 'Sí' if exento == 'S' else 'No']]
        t_fin = Table(fin_data, colWidths=[120, 360])
        t_fin.setStyle(TableStyle([('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 10), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]))
        elements.append(t_fin)
        elements.append(Spacer(1, 30))
        elements.append(Paragraph('_________________________&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;_________________________', styles['Normal']))
        elements.append(Paragraph('Firma del Beneficiario&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Firma del Responsable', styles['Normal']))
        doc.build(elements)
        return _pdf_payload(buf, f'contrato_comodato_{folio_com}.pdf')
    except (NotFoundError, ValidationError, InternalError):
        raise
    except Exception:
        logger.exception('Error al generar contrato de comodato PDF')
        raise InternalError(_MSG_ERROR_INTERNO)

def _exportar_beneficiarios_excel(genero: str | None=None, estado: str | None=None, membresia_estatus: str | None=None, busqueda: str | None=None, _current_user: CurrentUser | None = None):
    """Exportar lista filtrada de beneficiarios a Excel (RF-RB-07)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    try:
        with get_db() as conn:
            cur = conn.cursor()
            conditions = ["p.ACTIVO = 'S'", "p.ESTATUS_REGISTRO = 'APROBADO'"]
            params: dict = {}
            if genero:
                conditions.append('p.GENERO = :genero')
                params['genero'] = genero
            if estado:
                conditions.append('p.ESTADO = :estado')
                params['estado'] = estado
            if membresia_estatus:
                conditions.append('p.MEMBRESIA_ESTATUS = :membresia_estatus')
                params['membresia_estatus'] = membresia_estatus
            if busqueda:
                conditions.append('(LOWER(p.NOMBRE) LIKE :busqueda OR LOWER(p.APELLIDO_PATERNO) LIKE :busqueda OR LOWER(p.FOLIO) LIKE :busqueda)')
                params['busqueda'] = f'%{busqueda.lower()}%'
            where = ' AND '.join(conditions)
            cur.execute(f'SELECT p.FOLIO, p.NOMBRE, p.APELLIDO_PATERNO, p.APELLIDO_MATERNO,\n                           p.GENERO, p.FECHA_NACIMIENTO, p.ESTADO, p.CIUDAD,\n                           p.MEMBRESIA_ESTATUS, p.TIPO_CUOTA, p.FECHA_ALTA\n                    FROM PACIENTE p WHERE {where}\n                    ORDER BY p.ID_PACIENTE', params)
            rows = rows_to_dicts(cur)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Beneficiarios'
        headers = ['Folio', 'Nombre', 'Apellido Paterno', 'Apellido Materno', _COL_GENERO, 'Fecha Nacimiento', 'Estado', 'Ciudad', 'Membresía', 'Tipo Cuota', 'Fecha Alta']
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=_strip(row.get('folio')))
            ws.cell(row=row_idx, column=2, value=_strip(row.get('nombre')))
            ws.cell(row=row_idx, column=3, value=_strip(row.get('apellido_paterno')))
            ws.cell(row=row_idx, column=4, value=_strip(row.get('apellido_materno')))
            ws.cell(row=row_idx, column=5, value=_strip(row.get('genero')))
            ws.cell(row=row_idx, column=6, value=_date_str(row.get('fecha_nacimiento')))
            ws.cell(row=row_idx, column=7, value=_strip(row.get('estado')))
            ws.cell(row=row_idx, column=8, value=_strip(row.get('ciudad')))
            ws.cell(row=row_idx, column=9, value=_strip(row.get('membresia_estatus')))
            ws.cell(row=row_idx, column=10, value=_strip(row.get('tipo_cuota')))
            ws.cell(row=row_idx, column=11, value=_date_str(row.get('fecha_alta')))
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
        buf = io.BytesIO()
        wb.save(buf)
        return _excel_payload(buf, f'beneficiarios_{datetime.now().strftime('%Y%m%d')}.xlsx')
    except Exception:
        logger.exception('Error al generar Excel de beneficiarios')
        raise InternalError(_MSG_ERROR_INTERNO)

def _excel_write_headers(sheet, headers, header_fill, header_font) -> None:
    """Write a styled header row to an openpyxl worksheet."""
    from openpyxl.styles import Alignment
    for ci, h in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


def _excel_autosize_columns(wb) -> None:
    """Auto-fit all column widths (max 40) across every sheet in the workbook."""
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)


def _excel_fill_servicios_sheet(ws, fecha_inicio, fecha_fin, header_fill, header_font, current_user) -> None:
    from app.infrastructure.reportes.repository import reporte_servicios_por_tipo
    ws.title = 'Servicios por Tipo'
    data = reporte_servicios_por_tipo(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)
    _excel_write_headers(ws, ['Servicio', 'Cantidad', 'Monto'], header_fill, header_font)
    for i, (label, val) in enumerate(zip(data['labels'], data['values']), 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=3, value=data['montos'][i - 2])


def _excel_fill_pagos_sheet(ws, fecha_inicio, fecha_fin, header_fill, header_font, current_user) -> None:
    from app.infrastructure.reportes.repository import reporte_pagos_exentos
    ws.title = 'Pagos Exentos vs Cuotas'
    data = reporte_pagos_exentos(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)
    _excel_write_headers(ws, ['Concepto', 'Cantidad', 'Monto'], header_fill, header_font)
    ws.cell(row=2, column=1, value='Exentos')
    ws.cell(row=2, column=2, value=data['total_exentos'])
    ws.cell(row=2, column=3, value=data['monto_exentos'])
    ws.cell(row=3, column=1, value='Cuotas')
    ws.cell(row=3, column=2, value=data['total_cuotas'])
    ws.cell(row=3, column=3, value=data['monto_cuotas'])


def _excel_fill_consolidado_sheet(ws, mes, anio, header_fill, header_font, current_user) -> None:
    from app.infrastructure.reportes.repository import reporte_consolidado_mensual
    ws.title = 'Consolidado Mensual'
    data = reporte_consolidado_mensual(mes=mes, anio=anio, current_user=current_user)
    _excel_write_headers(ws, [_COL_METRICA, 'Valor'], header_fill, header_font)
    metrics: list = [
        ('Mes/Año', f'{data["mes"]}/{data["anio"]}'),
        (_COL_PAC_ATENDIDOS, data['pacientes_atendidos']),
        ('Total Servicios', data['total_servicios']),
        ('Monto Servicios', data['monto_servicios']),
        ('Total Ventas', data['total_ventas']),
        ('Monto Ventas', data['monto_ventas']),
    ]
    for k, v in data.get('citas_por_estatus', {}).items():
        metrics.append((f'Citas {k}', v))
    for k, v in data.get('por_genero', {}).items():
        metrics.append((f'Género {k}', v))
    for i, (label, val) in enumerate(metrics, 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)


def _excel_fill_resumen_sheet(ws, fecha_inicio, fecha_fin, header_fill, header_font, current_user) -> None:
    from app.infrastructure.reportes.repository import reporte_resumen
    ws.title = 'Resumen'
    data = reporte_resumen(genero=None, estado=None, tipo_espina=None,
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, current_user=current_user)
    _excel_write_headers(ws, [_COL_METRICA, 'Valor'], header_fill, header_font)
    metrics: list = [
        (_COL_TOTAL_PACIENTES, data['total_pacientes']),
        ('Activos', data['activos']),
        ('Inactivos', data['inactivos']),
        (_COL_EDAD_PROMEDIO, data['edad_promedio']),
        ('Estados Representados', data['estados_representados']),
    ]
    for k, v in data.get('por_genero', {}).items():
        metrics.append((f'Género: {k}', v))
    for i, (label, val) in enumerate(metrics, 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)


def _excel_write_all_report_sheets(wb, header_fill, header_font, fecha_inicio, fecha_fin, current_user) -> None:
    """Build the multi-sheet 'all' workbook: one sheet per standard report type."""
    from app.infrastructure.reportes.repository import (
        reporte_resumen, reporte_por_genero, reporte_por_etapa_vida,
        reporte_por_estado, reporte_por_tipo_espina,
        reporte_servicios_por_tipo, reporte_pagos_exentos,
    )
    ckw = {'current_user': current_user}
    bkw = {'genero': None, 'estado': None, 'tipo_espina': None,
           'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, **ckw}

    ws1 = wb.create_sheet('Resumen')
    d = reporte_resumen(**bkw)
    _excel_write_headers(ws1, [_COL_METRICA, 'Valor'], header_fill, header_font)
    items: list = [
        (_COL_TOTAL_PACIENTES, d['total_pacientes']), ('Activos', d['activos']),
        ('Inactivos', d['inactivos']), (_COL_EDAD_PROMEDIO, d['edad_promedio']),
        ('Estados Representados', d['estados_representados']),
    ]
    for k, v in d.get('por_genero', {}).items():
        items.append((f'Género: {k}', v))
    for k, v in d.get('por_tipo_espina', {}).items():
        items.append((f'Tipo Espina: {k}', v))
    for i, (lb, vl) in enumerate(items, 2):
        ws1.cell(row=i, column=1, value=lb)
        ws1.cell(row=i, column=2, value=vl)

    ws2 = wb.create_sheet('Por Género')
    d2 = reporte_por_genero(**bkw)
    _excel_write_headers(ws2, [_COL_GENERO, 'Cantidad'], header_fill, header_font)
    for i, (lb, vl) in enumerate(zip(d2['labels'], d2['values']), 2):
        ws2.cell(row=i, column=1, value=lb)
        ws2.cell(row=i, column=2, value=vl)

    ws3 = wb.create_sheet('Por Etapa de Vida')
    d3 = reporte_por_etapa_vida(**bkw)
    _excel_write_headers(ws3, [_COL_ETAPA_VIDA, 'Cantidad'], header_fill, header_font)
    for i, (lb, vl) in enumerate(zip(d3['labels'], d3['values']), 2):
        ws3.cell(row=i, column=1, value=lb)
        ws3.cell(row=i, column=2, value=vl)

    ws4 = wb.create_sheet('Por Estado')
    d4 = reporte_por_estado(**bkw)
    _excel_write_headers(ws4, ['Estado', 'Cantidad'], header_fill, header_font)
    for i, (lb, vl) in enumerate(zip(d4['labels'], d4['values']), 2):
        ws4.cell(row=i, column=1, value=lb)
        ws4.cell(row=i, column=2, value=vl)

    ws5 = wb.create_sheet('Por Tipo de Espina')
    d5 = reporte_por_tipo_espina(**bkw)
    _excel_write_headers(ws5, ['Tipo de Espina', 'Cantidad'], header_fill, header_font)
    for i, (lb, vl) in enumerate(zip(d5['labels'], d5['values']), 2):
        ws5.cell(row=i, column=1, value=lb)
        ws5.cell(row=i, column=2, value=vl)

    ws6 = wb.create_sheet('Servicios por Tipo')
    d6 = reporte_servicios_por_tipo(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, **ckw)
    _excel_write_headers(ws6, ['Servicio', 'Cantidad', 'Monto'], header_fill, header_font)
    for i, (lb, vl) in enumerate(zip(d6['labels'], d6['values']), 2):
        ws6.cell(row=i, column=1, value=lb)
        ws6.cell(row=i, column=2, value=vl)
        ws6.cell(row=i, column=3, value=d6['montos'][i - 2])

    ws7 = wb.create_sheet('Pagos Exentos')
    d7 = reporte_pagos_exentos(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, **ckw)
    _excel_write_headers(ws7, ['Concepto', 'Cantidad', 'Monto'], header_fill, header_font)
    ws7.cell(row=2, column=1, value='Exentos')
    ws7.cell(row=2, column=2, value=d7['total_exentos'])
    ws7.cell(row=2, column=3, value=d7['monto_exentos'])
    ws7.cell(row=3, column=1, value='Cuotas')
    ws7.cell(row=3, column=2, value=d7['total_cuotas'])
    ws7.cell(row=3, column=3, value=d7['monto_cuotas'])


def _exportar_reporte_excel(
    tipo: str = 'resumen',
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    mes: int | None = None,
    anio: int | None = None,
    current_user: CurrentUser | None = None,
):
    """Exportar datos de reportes a Excel (RF-ER-11)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    try:
        wb = Workbook()
        ws = wb.active
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        if tipo == 'all':
            wb.remove(ws)
            _excel_write_all_report_sheets(wb, header_fill, header_font, fecha_inicio, fecha_fin, current_user)
        elif tipo == 'servicios-por-tipo':
            _excel_fill_servicios_sheet(ws, fecha_inicio, fecha_fin, header_fill, header_font, current_user)
        elif tipo == 'pagos-exentos':
            _excel_fill_pagos_sheet(ws, fecha_inicio, fecha_fin, header_fill, header_font, current_user)
        elif tipo == 'consolidado-mensual':
            _excel_fill_consolidado_sheet(ws, mes, anio, header_fill, header_font, current_user)
        else:
            _excel_fill_resumen_sheet(ws, fecha_inicio, fecha_fin, header_fill, header_font, current_user)

        _excel_autosize_columns(wb)

        buf = io.BytesIO()
        wb.save(buf)
        filename = (
            f'reportes_completo_{datetime.now().strftime("%Y%m%d")}.xlsx'
            if tipo == 'all'
            else f'reporte_{tipo}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        return _excel_payload(buf, filename)
    except Exception:
        logger.exception('Error al generar Excel de reportes')
        raise InternalError(_MSG_ERROR_INTERNO)


class OracleExportacionesRepository(ExportacionesRepository):
    def exportar_reporte_pdf(self, tipo='resumen', genero=None, estado=None, tipo_espina=None, fecha_inicio=None, fecha_fin=None, mes=None, anio=None, current_user=None):
        return _exportar_reporte_pdf(tipo, genero, estado, tipo_espina, fecha_inicio, fecha_fin, mes, anio, current_user)

    def exportar_beneficiario_pdf(self, folio, current_user=None):
        return _exportar_beneficiario_pdf(folio, current_user)

    def exportar_credencial_pdf(self, folio, current_user=None):
        return _exportar_credencial_pdf(folio, current_user)

    def exportar_comprobante_cita(self, id_cita, current_user=None):
        return _exportar_comprobante_cita(id_cita, current_user)

    def exportar_contrato_comodato(self, id_comodato, current_user=None):
        return _exportar_contrato_comodato(id_comodato, current_user)

    def exportar_beneficiarios_excel(self, genero=None, estado=None, membresia_estatus=None, busqueda=None, current_user=None):
        return _exportar_beneficiarios_excel(genero, estado, membresia_estatus, busqueda, current_user)

    def exportar_reporte_excel(self, tipo='resumen', fecha_inicio=None, fecha_fin=None, mes=None, anio=None, current_user=None):
        return _exportar_reporte_excel(tipo, fecha_inicio, fecha_fin, mes, anio, current_user)
