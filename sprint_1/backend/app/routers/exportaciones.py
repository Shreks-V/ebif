"""
Exportaciones — PDF y Excel.

Cubre:
  RF-ER-05  Reportes en formato PDF
  RF-ER-06  Reporte general de un paciente en PDF
  RF-ER-11  Exportar tablas a Excel
  RF-RB-06  Credencial de beneficiario en PDF
  RF-RB-07  Exportar lista filtrada de beneficiarios a Excel
  RF-SO-10  Comprobante de servicio / cita en PDF
  RF-PS-05  Contrato de comodato en PDF
"""
import io
import logging
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse

from app.core.security import get_current_user
from app.core.database import get_db, rows_to_dicts, row_to_dict
from app.core.crypto import decrypt_row, PACIENTE_ENCRYPTED_FIELDS

logger = logging.getLogger(__name__)

router = APIRouter()


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════


def _strip(val):
    return val.strip() if isinstance(val, str) else val


def _date_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    return str(val)


def _pdf_response(buffer: io.BytesIO, filename: str) -> StreamingResponse:
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _excel_response(buffer: io.BytesIO, filename: str) -> StreamingResponse:
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════
#  RF-ER-05 / RF-ER-06 — Reporte general PDF
# ══════════════════════════════════════════════════════════════


@router.get("/reportes/pdf")
def exportar_reporte_pdf(
    tipo: str = Query("resumen", description="resumen | por-genero | por-etapa-vida | por-estado | por-tipo-espina"),
    genero: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    tipo_espina: Optional[int] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Generar reporte estadístico en PDF (RF-ER-05)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    try:
        # Import the report logic from reportes router
        from app.routers.reportes import (
            reporte_resumen,
            reporte_por_genero,
            reporte_por_etapa_vida,
            reporte_por_estado,
            reporte_por_tipo_espina,
        )

        # Build filter kwargs
        kwargs = {
            "genero": genero, "estado": estado, "tipo_espina": tipo_espina,
            "fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin,
            "current_user": current_user,
        }

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph("Asociación de Espina Bífida — Reporte", styles["Title"]))
        elements.append(Paragraph(f"Tipo: {tipo} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        if tipo == "resumen":
            data = reporte_resumen(**kwargs)
            table_data = [
                ["Métrica", "Valor"],
                ["Total Pacientes", str(data["total_pacientes"])],
                ["Activos", str(data["activos"])],
                ["Inactivos", str(data["inactivos"])],
                ["Edad Promedio", str(data["edad_promedio"])],
                ["Estados Representados", str(data["estados_representados"])],
            ]
            for gen, cnt in data.get("por_genero", {}).items():
                table_data.append([f"Género: {gen}", str(cnt)])
            for tip, cnt in data.get("por_tipo_espina", {}).items():
                table_data.append([f"Tipo Espina: {tip}", str(cnt)])
        else:
            report_funcs = {
                "por-genero": reporte_por_genero,
                "por-etapa-vida": reporte_por_etapa_vida,
                "por-estado": reporte_por_estado,
                "por-tipo-espina": reporte_por_tipo_espina,
            }
            func = report_funcs.get(tipo)
            if not func:
                raise HTTPException(status_code=400, detail=f"Tipo de reporte no válido: {tipo}")
            data = func(**kwargs)
            table_data = [["Categoría", "Cantidad"]]
            for label, value in zip(data["labels"], data["values"]):
                table_data.append([str(label), str(value)])
            table_data.append(["Total", str(data["total"])])

        t = Table(table_data, colWidths=[300, 150])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        doc.build(elements)

        return _pdf_response(buf, f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d')}.pdf")

    except HTTPException:
        raise
    except Exception:
        logger.exception("Error al generar PDF de reporte")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ══════════════════════════════════════════════════════════════
#  RF-ER-06 — Reporte general de un beneficiario en PDF
# ══════════════════════════════════════════════════════════════


@router.get("/beneficiario/{folio}/pdf")
def exportar_beneficiario_pdf(
    folio: str,
    current_user: dict = Depends(get_current_user),
):
    """Generar reporte PDF de un beneficiario con sus datos y documentos (RF-ER-06)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM PACIENTE WHERE FOLIO = :folio", {"folio": folio})
            paciente = row_to_dict(cur)
            if not paciente:
                raise HTTPException(status_code=404, detail="Beneficiario no encontrado")
            paciente = decrypt_row(paciente, PACIENTE_ENCRYPTED_FIELDS)

            nombre = f"{paciente['nombre']} {paciente['apellido_paterno']} {paciente.get('apellido_materno') or ''}".strip()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Asociación de Espina Bífida", styles["Title"]))
        elements.append(Paragraph(f"Expediente del Beneficiario — {nombre}", styles["Heading2"]))
        elements.append(Paragraph(f"Folio: {folio} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        fields = [
            ("Nombre", nombre),
            ("CURP", paciente.get("curp") or ""),
            ("Género", _strip(paciente.get("genero")) or ""),
            ("Fecha de Nacimiento", _date_str(paciente.get("fecha_nacimiento"))),
            ("Tipo de Sangre", paciente.get("tipo_sangre") or ""),
            ("Usa Válvula", "Sí" if _strip(paciente.get("usa_valvula")) == "S" else "No"),
            ("Dirección", paciente.get("direccion") or ""),
            ("Colonia", paciente.get("colonia") or ""),
            ("Ciudad", paciente.get("ciudad") or ""),
            ("Estado", paciente.get("estado") or ""),
            ("C.P.", paciente.get("codigo_postal") or ""),
            ("Teléfono Casa", paciente.get("telefono_casa") or ""),
            ("Teléfono Celular", paciente.get("telefono_celular") or ""),
            ("Correo", paciente.get("correo_electronico") or ""),
            ("Contacto Emergencia", paciente.get("en_emergencia_avisar_a") or ""),
            ("Tel. Emergencia", paciente.get("telefono_emergencia") or ""),
            ("Membresía", _strip(paciente.get("membresia_estatus")) or ""),
            ("Tipo Cuota", _strip(paciente.get("tipo_cuota")) or ""),
            ("Fecha Alta", _date_str(paciente.get("fecha_alta"))),
        ]

        table_data = [["Campo", "Valor"]]
        for label, val in fields:
            table_data.append([label, str(val)])

        t = Table(table_data, colWidths=[180, 300])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)

        return _pdf_response(buf, f"expediente_{folio}_{datetime.now().strftime('%Y%m%d')}.pdf")

    except HTTPException:
        raise
    except Exception:
        logger.exception("Error al generar PDF del beneficiario")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ══════════════════════════════════════════════════════════════
#  RF-RB-06 — Credencial del beneficiario en PDF
# ══════════════════════════════════════════════════════════════


@router.get("/beneficiario/{folio}/credencial")
def exportar_credencial_pdf(
    folio: str,
    current_user: dict = Depends(get_current_user),
):
    """Generar credencial del beneficiario en PDF (RF-RB-06)."""
    from reportlab.lib.pagesizes import landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas

    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM PACIENTE WHERE FOLIO = :folio", {"folio": folio})
            paciente = row_to_dict(cur)
            if not paciente:
                raise HTTPException(status_code=404, detail="Beneficiario no encontrado")
            paciente = decrypt_row(paciente, PACIENTE_ENCRYPTED_FIELDS)

            # Fetch tipos espina
            cur.execute(
                """SELECT te.NOMBRE FROM PACIENTE_TIPO_ESPINA pte
                   JOIN TIPO_ESPINA_BIFIDA te ON te.ID_TIPO_ESPINA = pte.ID_TIPO_ESPINA
                   WHERE pte.ID_PACIENTE = :id""",
                {"id": paciente["id_paciente"]},
            )
            tipos = [r[0].strip() for r in cur.fetchall()]

        nombre = f"{paciente['nombre']} {paciente['apellido_paterno']} {paciente.get('apellido_materno') or ''}".strip()

        buf = io.BytesIO()
        # Credit card size ~ 8.56 x 5.4 cm, scaled up for readability
        w, h = 10 * cm, 6.5 * cm
        c = pdf_canvas.Canvas(buf, pagesize=(w, h))

        # Background
        c.setFillColor(colors.HexColor("#1e3a5f"))
        c.rect(0, 0, w, h, fill=1)

        # Header bar
        c.setFillColor(colors.HexColor("#2d5f8a"))
        c.rect(0, h - 1.6 * cm, w, 1.6 * cm, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(w / 2, h - 1.1 * cm, "Asociación de Espina Bífida")
        c.setFont("Helvetica", 7)
        c.drawCentredString(w / 2, h - 1.4 * cm, "Credencial de Beneficiario")

        # Data
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 9)
        y = h - 2.3 * cm
        c.drawString(0.5 * cm, y, f"Nombre: {nombre}")
        y -= 0.55 * cm
        c.setFont("Helvetica", 8)
        c.drawString(0.5 * cm, y, f"Folio: {folio}")
        y -= 0.5 * cm
        c.drawString(0.5 * cm, y, f"CURP: {paciente.get('curp') or 'N/A'}")
        y -= 0.5 * cm
        c.drawString(0.5 * cm, y, f"Tipo de Sangre: {paciente.get('tipo_sangre') or 'N/A'}")
        y -= 0.5 * cm
        c.drawString(0.5 * cm, y, f"Diagnóstico: {', '.join(tipos) if tipos else 'N/A'}")
        y -= 0.5 * cm
        c.drawString(0.5 * cm, y, f"Membresía: {_strip(paciente.get('membresia_estatus')) or 'N/A'}")

        c.save()

        return _pdf_response(buf, f"credencial_{folio}.pdf")

    except HTTPException:
        raise
    except Exception:
        logger.exception("Error al generar credencial PDF")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ══════════════════════════════════════════════════════════════
#  RF-SO-10 — Comprobante de cita / servicio en PDF
# ══════════════════════════════════════════════════════════════


@router.get("/cita/{id_cita}/comprobante")
def exportar_comprobante_cita(
    id_cita: int,
    current_user: dict = Depends(get_current_user),
):
    """Generar comprobante PDF de una cita con sus servicios (RF-SO-10)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute(
                """SELECT c.ID_CITA, c.FECHA_HORA, c.ESTATUS, c.NOTAS,
                          p.NOMBRE || ' ' || p.APELLIDO_PATERNO || ' ' || NVL(p.APELLIDO_MATERNO, '') AS nombre_paciente,
                          p.FOLIO AS folio_paciente
                   FROM CITA c
                   JOIN PACIENTE p ON p.ID_PACIENTE = c.ID_PACIENTE
                   WHERE c.ID_CITA = :id_cita""",
                {"id_cita": id_cita},
            )
            cita = row_to_dict(cur)
            if not cita:
                raise HTTPException(status_code=404, detail="Cita no encontrada")

            # Servicios
            cur.execute(
                """SELECT s.NOMBRE, d.CANTIDAD, d.MONTO_PAGADO, d.CANCELADO
                   FROM DETALLE_CITA_SERVICIO d
                   JOIN SERVICIO s ON s.ID_SERVICIO = d.ID_SERVICIO
                   WHERE d.ID_CITA = :id_cita""",
                {"id_cita": id_cita},
            )
            servicios = rows_to_dicts(cur)

            # Doctores
            cur.execute(
                """SELECT d.NOMBRE || ' ' || d.APELLIDO_PATERNO AS nombre_doctor,
                          d.ESPECIALIDAD, cd.ROL_DOCTOR
                   FROM CITA_DOCTOR cd
                   JOIN DOCTOR d ON d.ID_DOCTOR = cd.ID_DOCTOR
                   WHERE cd.ID_CITA = :id_cita""",
                {"id_cita": id_cita},
            )
            doctores = rows_to_dicts(cur)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Asociación de Espina Bífida", styles["Title"]))
        elements.append(Paragraph("Comprobante de Servicio", styles["Heading2"]))
        elements.append(Spacer(1, 10))

        # Cita info
        info_data = [
            ["Paciente", cita["nombre_paciente"]],
            ["Folio Paciente", cita["folio_paciente"]],
            ["Fecha", _date_str(cita["fecha_hora"])],
            ["Estatus", _strip(cita["estatus"])],
            ["Notas", cita.get("notas") or ""],
        ]
        if doctores:
            for doc_row in doctores:
                info_data.append(["Doctor", f"{doc_row['nombre_doctor']} — {_strip(doc_row.get('especialidad') or '')}"])

        t_info = Table(info_data, colWidths=[120, 360])
        t_info.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t_info)
        elements.append(Spacer(1, 15))

        # Servicios table
        elements.append(Paragraph("Servicios", styles["Heading3"]))
        svc_data = [["Servicio", "Cantidad", "Monto", "Cancelado"]]
        total_monto = 0.0
        for s in servicios:
            monto = float(s.get("monto_pagado") or 0)
            total_monto += monto
            svc_data.append([
                _strip(s["nombre"]),
                str(s.get("cantidad", 1)),
                f"${monto:,.2f}",
                "Sí" if _strip(s.get("cancelado")) == "S" else "No",
            ])
        svc_data.append(["", "", f"Total: ${total_monto:,.2f}", ""])

        t_svc = Table(svc_data, colWidths=[200, 80, 100, 100])
        t_svc.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t_svc)

        doc.build(elements)
        return _pdf_response(buf, f"comprobante_cita_{id_cita}.pdf")

    except HTTPException:
        raise
    except Exception:
        logger.exception("Error al generar comprobante PDF")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ══════════════════════════════════════════════════════════════
#  RF-PS-05 — Contrato de comodato en PDF
# ══════════════════════════════════════════════════════════════


@router.get("/comodato/{id_comodato}/contrato")
def exportar_contrato_comodato(
    id_comodato: int,
    current_user: dict = Depends(get_current_user),
):
    """Generar contrato de comodato en PDF (RF-PS-05)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute(
                """SELECT c.*, p.NOMBRE || ' ' || p.APELLIDO_PATERNO || ' ' || NVL(p.APELLIDO_MATERNO, '') AS nombre_paciente,
                          p.FOLIO AS folio_paciente, pr.NOMBRE AS nombre_equipo,
                          eq.NUMERO_SERIE, eq.MARCA, eq.MODELO
                   FROM COMODATO c
                   JOIN PACIENTE p ON p.ID_PACIENTE = c.ID_PACIENTE
                   JOIN PRODUCTO pr ON pr.ID_PRODUCTO = c.ID_EQUIPO
                   LEFT JOIN EQUIPO_MEDICO eq ON eq.ID_PRODUCTO = c.ID_EQUIPO
                   WHERE c.ID_COMODATO = :id""",
                {"id": id_comodato},
            )
            com = row_to_dict(cur)
            if not com:
                raise HTTPException(status_code=404, detail="Comodato no encontrado")

            # Get patient details for the contract
            cur.execute("SELECT * FROM PACIENTE WHERE ID_PACIENTE = :id", {"id": com["id_paciente"]})
            paciente = row_to_dict(cur)
            paciente = decrypt_row(paciente, PACIENTE_ENCRYPTED_FIELDS)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=50, bottomMargin=50)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Asociación de Espina Bífida", styles["Title"]))
        elements.append(Paragraph("CONTRATO DE COMODATO", styles["Heading2"]))
        elements.append(Spacer(1, 15))

        folio_com = _strip(com.get("folio_comodato")) or str(id_comodato)
        elements.append(Paragraph(f"<b>Folio:</b> {folio_com}", styles["Normal"]))
        elements.append(Paragraph(
            f"<b>Fecha:</b> {_date_str(com.get('fecha_prestamo'))}",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 10))

        # Beneficiary
        elements.append(Paragraph("<b>DATOS DEL BENEFICIARIO</b>", styles["Heading3"]))
        ben_data = [
            ["Nombre", com["nombre_paciente"]],
            ["Folio", com["folio_paciente"]],
            ["Dirección", paciente.get("direccion") or ""],
            ["Teléfono", paciente.get("telefono_celular") or paciente.get("telefono_casa") or ""],
        ]
        t_ben = Table(ben_data, colWidths=[120, 360])
        t_ben.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t_ben)
        elements.append(Spacer(1, 10))

        # Equipment
        elements.append(Paragraph("<b>EQUIPO EN PRÉSTAMO</b>", styles["Heading3"]))
        eq_data = [
            ["Equipo", _strip(com.get("nombre_equipo")) or ""],
            ["Número de Serie", _strip(com.get("numero_serie")) or "N/A"],
            ["Marca", _strip(com.get("marca")) or "N/A"],
            ["Modelo", _strip(com.get("modelo")) or "N/A"],
        ]
        t_eq = Table(eq_data, colWidths=[120, 360])
        t_eq.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t_eq)
        elements.append(Spacer(1, 10))

        # Financial
        elements.append(Paragraph("<b>CONDICIONES ECONÓMICAS</b>", styles["Heading3"]))
        monto_total = float(com.get("monto_total") or 0)
        monto_pagado = float(com.get("monto_pagado") or 0)
        saldo = float(com.get("saldo_pendiente") or 0)
        exento = _strip(com.get("exento_pago"))
        fin_data = [
            ["Monto Total", f"${monto_total:,.2f}"],
            ["Monto Pagado", f"${monto_pagado:,.2f}"],
            ["Saldo Pendiente", f"${saldo:,.2f}"],
            ["Exento de Pago", "Sí" if exento == "S" else "No"],
        ]
        t_fin = Table(fin_data, colWidths=[120, 360])
        t_fin.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t_fin)
        elements.append(Spacer(1, 30))

        # Signature lines
        elements.append(Paragraph("_________________________&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;_________________________", styles["Normal"]))
        elements.append(Paragraph("Firma del Beneficiario&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Firma del Responsable", styles["Normal"]))

        doc.build(elements)
        return _pdf_response(buf, f"contrato_comodato_{folio_com}.pdf")

    except HTTPException:
        raise
    except Exception:
        logger.exception("Error al generar contrato de comodato PDF")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ══════════════════════════════════════════════════════════════
#  RF-RB-07 / RF-ER-11 — Exportar a Excel
# ══════════════════════════════════════════════════════════════


@router.get("/beneficiarios/excel")
def exportar_beneficiarios_excel(
    genero: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    membresia_estatus: Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Exportar lista filtrada de beneficiarios a Excel (RF-RB-07)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        with get_db() as conn:
            cur = conn.cursor()
            conditions = ["p.ACTIVO = 'S'", "p.ESTATUS_REGISTRO = 'APROBADO'"]
            params: dict = {}

            if genero:
                conditions.append("p.GENERO = :genero")
                params["genero"] = genero
            if estado:
                conditions.append("p.ESTADO = :estado")
                params["estado"] = estado
            if membresia_estatus:
                conditions.append("p.MEMBRESIA_ESTATUS = :membresia_estatus")
                params["membresia_estatus"] = membresia_estatus
            if busqueda:
                conditions.append(
                    "(LOWER(p.NOMBRE) LIKE :busqueda OR LOWER(p.APELLIDO_PATERNO) LIKE :busqueda "
                    "OR LOWER(p.FOLIO) LIKE :busqueda)"
                )
                params["busqueda"] = f"%{busqueda.lower()}%"

            where = " AND ".join(conditions)
            cur.execute(
                f"""SELECT p.FOLIO, p.NOMBRE, p.APELLIDO_PATERNO, p.APELLIDO_MATERNO,
                           p.GENERO, p.FECHA_NACIMIENTO, p.ESTADO, p.CIUDAD,
                           p.MEMBRESIA_ESTATUS, p.TIPO_CUOTA, p.FECHA_ALTA
                    FROM PACIENTE p WHERE {where}
                    ORDER BY p.ID_PACIENTE""",
                params,
            )
            rows = rows_to_dicts(cur)

        wb = Workbook()
        ws = wb.active
        ws.title = "Beneficiarios"

        headers = ["Folio", "Nombre", "Apellido Paterno", "Apellido Materno",
                    "Género", "Fecha Nacimiento", "Estado", "Ciudad",
                    "Membresía", "Tipo Cuota", "Fecha Alta"]

        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=_strip(row.get("folio")))
            ws.cell(row=row_idx, column=2, value=_strip(row.get("nombre")))
            ws.cell(row=row_idx, column=3, value=_strip(row.get("apellido_paterno")))
            ws.cell(row=row_idx, column=4, value=_strip(row.get("apellido_materno")))
            ws.cell(row=row_idx, column=5, value=_strip(row.get("genero")))
            ws.cell(row=row_idx, column=6, value=_date_str(row.get("fecha_nacimiento")))
            ws.cell(row=row_idx, column=7, value=_strip(row.get("estado")))
            ws.cell(row=row_idx, column=8, value=_strip(row.get("ciudad")))
            ws.cell(row=row_idx, column=9, value=_strip(row.get("membresia_estatus")))
            ws.cell(row=row_idx, column=10, value=_strip(row.get("tipo_cuota")))
            ws.cell(row=row_idx, column=11, value=_date_str(row.get("fecha_alta")))

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        buf = io.BytesIO()
        wb.save(buf)
        return _excel_response(buf, f"beneficiarios_{datetime.now().strftime('%Y%m%d')}.xlsx")

    except Exception:
        logger.exception("Error al generar Excel de beneficiarios")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.get("/reportes/excel")
def exportar_reporte_excel(
    tipo: str = Query("resumen", description="resumen | servicios-por-tipo | pagos-exentos | consolidado-mensual"),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    mes: Optional[int] = Query(None),
    anio: Optional[int] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Exportar datos de reportes a Excel (RF-ER-11)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.routers.reportes import (
        reporte_resumen,
        reporte_por_genero,
        reporte_por_etapa_vida,
        reporte_por_estado,
        reporte_servicios_por_tipo,
        reporte_pagos_exentos,
        reporte_consolidado_mensual,
    )

    try:
        wb = Workbook()
        ws = wb.active
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        def write_headers(sheet, headers):
            for ci, h in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        common_kwargs = {"current_user": current_user}

        if tipo == "servicios-por-tipo":
            ws.title = "Servicios por Tipo"
            data = reporte_servicios_por_tipo(
                fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, **common_kwargs
            )
            write_headers(ws, ["Servicio", "Cantidad", "Monto"])
            for i, (label, val) in enumerate(zip(data["labels"], data["values"]), 2):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=val)
                ws.cell(row=i, column=3, value=data["montos"][i - 2])

        elif tipo == "pagos-exentos":
            ws.title = "Pagos Exentos vs Cuotas"
            data = reporte_pagos_exentos(
                fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, **common_kwargs
            )
            write_headers(ws, ["Concepto", "Cantidad", "Monto"])
            ws.cell(row=2, column=1, value="Exentos")
            ws.cell(row=2, column=2, value=data["total_exentos"])
            ws.cell(row=2, column=3, value=data["monto_exentos"])
            ws.cell(row=3, column=1, value="Cuotas")
            ws.cell(row=3, column=2, value=data["total_cuotas"])
            ws.cell(row=3, column=3, value=data["monto_cuotas"])

        elif tipo == "consolidado-mensual":
            ws.title = "Consolidado Mensual"
            data = reporte_consolidado_mensual(mes=mes, anio=anio, **common_kwargs)
            write_headers(ws, ["Métrica", "Valor"])
            metrics = [
                ("Mes/Año", f"{data['mes']}/{data['anio']}"),
                ("Pacientes Atendidos", data["pacientes_atendidos"]),
                ("Total Servicios", data["total_servicios"]),
                ("Monto Servicios", data["monto_servicios"]),
                ("Total Ventas", data["total_ventas"]),
                ("Monto Ventas", data["monto_ventas"]),
            ]
            for k, v in data.get("citas_por_estatus", {}).items():
                metrics.append((f"Citas {k}", v))
            for k, v in data.get("por_genero", {}).items():
                metrics.append((f"Género {k}", v))
            for i, (label, val) in enumerate(metrics, 2):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=val)

        else:  # resumen
            ws.title = "Resumen"
            data = reporte_resumen(
                genero=None, estado=None, tipo_espina=None,
                fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                **common_kwargs,
            )
            write_headers(ws, ["Métrica", "Valor"])
            metrics = [
                ("Total Pacientes", data["total_pacientes"]),
                ("Activos", data["activos"]),
                ("Inactivos", data["inactivos"]),
                ("Edad Promedio", data["edad_promedio"]),
                ("Estados Representados", data["estados_representados"]),
            ]
            for k, v in data.get("por_genero", {}).items():
                metrics.append((f"Género: {k}", v))
            for i, (label, val) in enumerate(metrics, 2):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=val)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        buf = io.BytesIO()
        wb.save(buf)
        return _excel_response(buf, f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx")

    except Exception:
        logger.exception("Error al generar Excel de reportes")
        raise HTTPException(status_code=500, detail="Error interno del servidor")
